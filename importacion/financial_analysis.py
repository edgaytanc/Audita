import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font


def analizar_balance_general(archivo_excel_Anterior, archivo_excel_Actual): 
    # Leer archivos de Excel
    df_anterior = pd.read_excel(archivo_excel_Anterior, engine='openpyxl')
    df_actual = pd.read_excel(archivo_excel_Actual, engine='openpyxl')

    # Asegurarse de que ambos DataFrames tienen las mismas etiquetas de columnas
    assert set(df_anterior.columns.str.strip()) == set(df_actual.columns.str.strip()), "Las etiquetas de las columnas de los DataFrames no coinciden"

    # Convertir celdas no numéricas a NaN y luego reemplazarlas con 0
    df_anterior.iloc[:, 1:] = df_anterior.iloc[:, 1:].apply(pd.to_numeric, errors='coerce').fillna(0)
    df_actual.iloc[:, 1:] = df_actual.iloc[:, 1:].apply(pd.to_numeric, errors='coerce').fillna(0)

    # Crear DataFrame analisis_horizontal
    analisis_horizontal = pd.DataFrame()
    analisis_horizontal['CUENTA'] = df_anterior.iloc[:, 0]
    analisis_horizontal['SALDO ANTERIOR'] = df_anterior.iloc[:, 1]
    analisis_horizontal['SALDO ACTUAL'] = df_actual.iloc[:, 1]
    analisis_horizontal['VARIACIONES'] = analisis_horizontal['SALDO ACTUAL'] - analisis_horizontal['SALDO ANTERIOR']
    analisis_horizontal['PORCENTAJE'] = (analisis_horizontal['VARIACIONES'] / (analisis_horizontal['SALDO ACTUAL']) * 100)
    analisis_horizontal['ASEVERACIONES']=""
    analisis_horizontal['EVALUACIÓN DE RIESGO'] = pd.cut(analisis_horizontal['PORCENTAJE'],bins=[float('-inf'), 10, 20, float('inf')],labels=['Baja', 'Media', 'Alta'])
    return analisis_horizontal


def guardar_analisis_horizontal_excel(df, archivo_resultado):
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Aplicar negrita y tamaño de fuente 12 a la primera fila (encabezado)
    for celda in ws["1"]:
        celda.font = Font(bold=True, size=12)

    # Encuentra la fila que contiene 'TOTAL ACTIVO' o 'TOTAL PASIVO Y PATRIMONIO' y aplica el relleno celeste
    for row in ws.iter_rows():
        if row[0].value == 'TOTAL ACTIVO' or row[0].value == 'TOTAL PASIVO Y PATRIMONIO':
            celeste_fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")
            for celda in row:
                celda.fill = celeste_fill
    
    # Aplicar formato de moneda para Guatemala y porcentaje a las columnas correspondientes
    for row in ws.iter_rows(min_row=2):
        row[1].number_format = '#,##0.00;[Red]-#,##0.00'
        row[2].number_format = '#,##0.00;[Red]-#,##0.00'
        row[3].number_format = '#,##0.00;[Red]-#,##0.00'
        row[4].number_format = '#,##0.00'

    # Aplicar colores de fondo a la columna EVALUACION DE RIESGO según el valor
    for celda in ws['G']:
        if celda.value == 'Baja':
            celda.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type='solid')
        elif celda.value == 'Media':
            celda.fill = PatternFill(start_color="008000", end_color="008000", fill_type='solid')
        elif celda.value == 'Alta':
            celda.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type='solid')

    wb.save(archivo_resultado)

#PARA EL ANALISIS VERTICAL
def analizar_balance_vertical(archivo_excel):
    # Leer el archivo de Excel
    df = pd.read_excel(archivo_excel, engine='openpyxl')

    # Convertir celdas no numéricas a NaN y luego reemplazarlas con 0
    df.iloc[:, 1:] = df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce').fillna(0)

    # Crear un nuevo DataFrame para almacenar el análisis vertical
    analisis_vertical = pd.DataFrame()
    analisis_vertical['CUENTA'] = df['CUENTA']
    analisis_vertical['SALDO'] = df['SALDO']

    # Encontrar el índice de las filas 'TOTAL ACTIVO' y 'TOTAL PASIVO Y PATRIMONIO'
    total_activo_index = df[df['CUENTA'] == 'TOTAL ACTIVO'].index[0]
    total_pasivo_patrimonio_index = df[df['CUENTA'] == 'TOTAL PASIVO Y PATRIMONIO'].index[0]

    # Calcular el análisis vertical para activos y pasivos/patrimonio
    porcentaje = []
    for i, row in df.iterrows():
        if i <= total_activo_index:
            porcentaje.append(row['SALDO'] / df.at[total_activo_index, 'SALDO'] * 100)
        elif i > total_activo_index and i <= total_pasivo_patrimonio_index:
            porcentaje.append(row['SALDO'] / df.at[total_pasivo_patrimonio_index, 'SALDO'] * 100)
        else:
            porcentaje.append(None)

    analisis_vertical['PORCENTAJE'] = porcentaje

    # Devuelve los nombres de las cuentas, los saldos y los porcentajes
    return analisis_vertical['CUENTA'], analisis_vertical['SALDO'], analisis_vertical['PORCENTAJE']





def guardar_analisis_vertical_excel(cuentas_anterior, saldo_anterior, porcentaje_anterior, cuentas_actual, saldo_actual, porcentaje_actual, archivo_resultado):
    wb = Workbook()
    ws = wb.active

    # Ajusta el ancho de la columna A a 15.5
    ws.column_dimensions['A'].width = 15.5

    # Combina y centra las celdas A1, B1 y C1
    ws.merge_cells('A1:E1')
    titulo_celda = ws['A1']

    # Escribe el título "Análisis Vertical de Balance General" en la fila 1
    titulo = "Análisis Vertical de Balance General"
    titulo_celda = ws.cell(row=1, column=1)
    titulo_celda.value = titulo
    titulo_celda.font = Font(bold=True, size=13)

    # Escribe los encabezados
    ws.append(['CUENTA', 'SALDO ANTERIOR', 'PORCENTAJE ANTERIOR', 'SALDO ACTUAL', 'PORCENTAJE ACTUAL'])

    # Escribe los datos
    for cuenta_ant, saldo_ant, porcentaje_ant, cuenta_act, saldo_act, porcentaje_act in zip(cuentas_anterior, saldo_anterior, porcentaje_anterior, cuentas_actual, saldo_actual, porcentaje_actual):
        # Asegúrate de que las cuentas coinciden antes de escribir los datos
        assert cuenta_ant == cuenta_act, f"Las cuentas no coinciden: {cuenta_ant} != {cuenta_act}"
        ws.append([cuenta_ant, saldo_ant, porcentaje_ant, saldo_act, porcentaje_act])

    # Aplica formato numerico a las columnas
    for row in ws.iter_rows(min_row=3):
        for cell in row[1:]:
            cell.number_format = '#,##0.00;[Red]-#,##0.00'

    # Guarda el archivo de Excel
    wb.save(archivo_resultado)

