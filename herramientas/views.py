from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from datetime import datetime
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import rsa,padding
from .forms import NombramientoForm, ActividadForm
from .models import Firma, Nombramiento, Actividad
from django.contrib import messages
from django.shortcuts import get_object_or_404
from .forms import SeleccionarEntidadForm
from proyecto.models import Entidad
from .decorators import entidad_requerida
from django.contrib.auth import get_user_model
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from django.http import FileResponse
import tempfile
import openpyxl
from .models import Actividad
import os
from openpyxl.styles import Font, PatternFill, Color, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

# Obten la ruta absolua del directorio donde se encuetra este script
dir_path = os.path.dirname(os.path.realpath(__file__))

# Construye la ruta al archivo urls.json
urls_file_path =  os.path.join(dir_path, 'urls.json')

with open(urls_file_path) as f:
    urls = json.load(f)


@login_required
def firma_a_generar(request):
    current_user = request.user
    current_date = datetime.now().date()
    mensaje = ''
    if Firma.objects.filter(user=request.user).exists():
        mensaje = 'Ya se genero una Firma Digital'
        

    context = {
        'usuario': current_user,
        'fecha': current_date,
        'mensaje': mensaje
    }

    return render(request,'herramientas/firma_a_generar.html', context)

@login_required
def generar_firma(request):
    #obtenemos el usuario actual
    current_user = request.user

    #Genera la clave privada y publica del usuario
    private_key = rsa.generate_private_key(public_exponent=65537, key_size=2048)
    public_key = private_key.public_key()

    

    #Generar el mensaje a firmar
   
    message = bytes(request.POST['firma'],'utf-8')

    # Firmar el mensaje con la clave privada del usuario
    signature = private_key.sign(message, padding.PSS(mgf=padding.MGF1(hashes.SHA256()), salt_length=padding.PSS.MAX_LENGTH), hashes.SHA256())

    #Crea una instancia de Firma con la informacion de la firma digital
    firma_usuario = Firma(user=current_user,firma=signature,creado_en=datetime.now())

    #Guarda la instacia
    firma_usuario.save()

    return render(request, 'herramientas/firma_generada.html', {'firma':signature.hex()})

@login_required
@entidad_requerida
def crear_nombramiento(request):
    User = get_user_model()
    
    if request.method == 'POST':
        form = NombramientoForm(request.POST)
        if form.is_valid():
            nombramiento = form.save(commit=False)
            
            
            # Comprueba si el usuario ya tiene un nombramiento
            existing_nombramiento = Nombramiento.objects.filter(user=nombramiento.user).exists()

            if existing_nombramiento:
                messages.error(request, 'El usuario ya tiene un nombramiento creado.')
            else:
                nombramiento.save()
                messages.success(request, 'Nombramiento creado exitosamente')
                return redirect('/herramientas/nombramiento')
    else:
        form = NombramientoForm()

    context = {'form': form}
    return render(request, 'herramientas/nombramiento.html', context)



@login_required
@entidad_requerida
def listar_nombramientos(request):
    nombramientos = Nombramiento.objects.all()
    context = {'nombramientos': nombramientos}
    return render(request, 'herramientas/listar_nombramientos.html', context)



@login_required
@entidad_requerida
def editar_nombramiento(request, nombramiento_id):
    nombramiento = get_object_or_404(Nombramiento, id=nombramiento_id)
    if request.method == 'POST':
        form = NombramientoForm(request.POST, instance=nombramiento)
        if form.is_valid():
            form.save()
            messages.success(request, 'Nombramiento actualizado exitosamente')
            return redirect('/herramientas/listar_nombramientos')
    else:
        form = NombramientoForm(instance=nombramiento)

    context = {'form': form}
    return render(request, 'herramientas/editar_nombramiento.html', context)

@login_required
@entidad_requerida
def eliminar_nombramiento(request, nombramiento_id):
    nombramiento = get_object_or_404(Nombramiento, id=nombramiento_id)
    nombramiento.delete()
    messages.success(request, 'Nombramiento eliminado exitosamente')
    return redirect('/herramientas/listar_nombramientos')



@login_required
@entidad_requerida
def listar_actividades(request):
    actividades = Actividad.objects.all()
    return render(request, 'herramientas/listar_actividades.html', {'actividades': actividades})

@login_required
@entidad_requerida
def crear_actividad(request):
    if request.method == 'POST':
        form = ActividadForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_actividades')
    else:
        form = ActividadForm()
    return render(request, 'herramientas/crear_actividad.html', {'form': form})

@login_required
@entidad_requerida
def editar_actividad(request, actividad_id):
    actividad = Actividad.objects.get(id=actividad_id)
    if request.method == 'POST':
        form = ActividadForm(request.POST, instance=actividad)
        if form.is_valid():
            form.save()
            return redirect('listar_actividades')
    else:
        form = ActividadForm(instance=actividad)
    return render(request, 'herramientas/editar_actividad.html', {'form': form})

@login_required
@entidad_requerida
def eliminar_actividad(request, actividad_id):
    actividad = Actividad.objects.get(id=actividad_id)
    actividad.delete()
    return redirect('listar_actividades')

@login_required

def seleccionar_entidad(request):
    if request.method == 'POST':
        form = SeleccionarEntidadForm(request.POST)
        if form.is_valid():
            entidad_seleccionada = form.cleaned_data.get('entidad')
            request.session['entidad_seleccionada_id'] = entidad_seleccionada.id
            return redirect('index')
    else:
        form = SeleccionarEntidadForm()

    return render(request, 'herramientas/seleccionar_entidad.html', {'form': form})



# para la nueva Acividad
class ActividadListView(ListView):
    model = Actividad
    template_name = 'actividades/actividad_list.html'


class ActividadCreateView(CreateView):
    model = Actividad
    form_class = ActividadForm
    template_name = 'actividades/actividad_form.html'
    success_url = reverse_lazy('actividad_list')
    
    

class ActividadUpdateView(UpdateView):
    model = Actividad
    form_class = ActividadForm
    template_name = 'actividades/actividad_form.html'
    success_url = reverse_lazy('actividad_list')
    
    

class ActividadDeleteView(DeleteView):
    model = Actividad
    template_name = 'actividades/actividad_confirm_delete.html'
    success_url = reverse_lazy('actividad_list')


# crea el excel de las actividades
def some_view(request):
    # Create a workbook and add a worksheet to it
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Define the styles for the header
    font = Font(name='Calibri', color=Color("FFFFFF"), bold=True)  # White text, bold
    font2 = Font(name='Calibri', color=Color("000000"), bold=True)
    fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")  # Blue background
    alignment = Alignment(vertical="center")  # Center vertically

    # Combinar celdas
    sheet.merge_cells('H5:J5')
    sheet.merge_cells('H6:J6')
    sheet.merge_cells('H7:J7')
    sheet.merge_cells('H8:J8')

    # # Define tamano de la celda
    # sheet['H5'].width = 45
    # sheet['H6'].width = 45
    # sheet['H7'].width = 45
    # sheet['H8'].width = 45



    # Definir contenido de las celdas combinadas
    sheet['H5'] = 'ENTIDAD XXXXXXXXXXXXXXXXXXX'
    sheet['H6'] = 'AUDITORIA FINANCIERA'
    sheet['H7'] = 'DEL 01 DE ENERO AL 31 DE DICIEMBRE DE 2022'
    sheet['H8'] = 'CRONOGRAMA DE ACTIVIDADES'
    sheet['H8'].font = font2

    # Centrar texto de celdas combinadas
    alignment = Alignment(horizontal='center')
    sheet['H5'].alignment = alignment
    sheet['H6'].alignment = alignment
    sheet['H7'].alignment = alignment
    sheet['H8'].alignment = alignment

    # Define the border style
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Define the height of the header row in points
    header_row_height = 45.28  # 1.6 cm = ~45.28 points

    # Add headers to the worksheet
    headers = ['Actividad', 'Referencia', 'Nombramiento', 'Auditor', 'Fecha de inicio', 
               'Fecha de finalización', 'Estado Actual', 'Enero', 'Febrero', 'Marzo',
               'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre',
               'Noviembre', 'Diciembre', 'Total de días', 'Observaciones']
    
    for column in range(len(headers)):
        cell = sheet.cell(row=14, column=column + 1, value=headers[column])
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        # cell.border = thin_border

    # Set thi width of the columns
    sheet.column_dimensions[get_column_letter(headers.index('Actividad') + 1)].width = 50
    sheet.column_dimensions[get_column_letter(headers.index('Referencia') + 1)].width = 10
    sheet.column_dimensions[get_column_letter(headers.index('Nombramiento') + 1)].width = 20
    sheet.column_dimensions[get_column_letter(headers.index('Auditor') + 1)].width = 30
    sheet.column_dimensions[get_column_letter(headers.index('Fecha de inicio') + 1)].width = 22
    sheet.column_dimensions[get_column_letter(headers.index('Fecha de finalización') + 1)].width = 22
    sheet.column_dimensions[get_column_letter(headers.index('Estado Actual') + 1)].width = 15
    sheet.column_dimensions[get_column_letter(headers.index('Enero') + 1)].width = 12
    sheet.column_dimensions[get_column_letter(headers.index('Febrero') + 1)].width = 12
    sheet.column_dimensions[get_column_letter(headers.index('Marzo') + 1)].width = 12
    sheet.column_dimensions[get_column_letter(headers.index('Abril') + 1)].width = 12
    sheet.column_dimensions[get_column_letter(headers.index('Mayo') + 1)].width = 12
    sheet.column_dimensions[get_column_letter(headers.index('Junio') + 1)].width = 12
    sheet.column_dimensions[get_column_letter(headers.index('Julio') + 1)].width = 12
    sheet.column_dimensions[get_column_letter(headers.index('Agosto') + 1)].width = 12
    sheet.column_dimensions[get_column_letter(headers.index('Septiembre') + 1)].width = 12
    sheet.column_dimensions[get_column_letter(headers.index('Octubre') + 1)].width = 12
    sheet.column_dimensions[get_column_letter(headers.index('Noviembre') + 1)].width = 12
    sheet.column_dimensions[get_column_letter(headers.index('Diciembre') + 1)].width = 12
    sheet.column_dimensions[get_column_letter(headers.index('Total de días') + 1)].width = 15
    sheet.column_dimensions[get_column_letter(headers.index('Observaciones') + 1)].width = 50


    # Set the height of the header row
    sheet.row_dimensions[14].height = header_row_height

    # sheet.append(headers)

    # Add data to the worksheet
    actividades = Actividad.objects.all()
    row_idx = 15
    for actividad in actividades:
        row = [actividad.actividad, actividad.referencia, actividad.nombramiento, actividad.auditor, 
               actividad.fecha_inicio, actividad.fecha_finalizacion, actividad.estado_actual, 
               actividad.enero, actividad.febrero, actividad.marzo, actividad.abril, actividad.mayo,
               actividad.junio, actividad.julio, actividad.agosto, actividad.septiembre,
               actividad.octubre, actividad.noviembre, actividad.diciembre,actividad.total_dias, actividad.observaciones]
        for column in range(len(row)):
            cell = sheet.cell(row=row_idx,column=column+1, value=row[column])
            cell.border = thin_border
        row_idx +=1
        # sheet.append(row)

    # Encuentra la ultima fila de la tabla
    last_row = sheet.max_row + 3

    # Anade el texto en las siguientes tres filas en la columna B
    sheet['B' + str(last_row + 1)] = 'Firma: _________________________________'
    sheet['B' + str(last_row + 2)] = 'Elaboró: _______________________________'
    sheet['B' + str(last_row + 3)] = 'Fecha: _________________________________'

    sheet['O' + str(last_row + 1)] = 'Firma: _________________________________'
    sheet['O' + str(last_row + 2)] = 'Revisó: ________________________________'
    sheet['O' + str(last_row + 3)] = 'Fecha: _________________________________'

    # Save workbook to temporary file
    temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    workbook.save(temp.name)
    temp.close()
    
    # Create a response with it
    response = FileResponse(open(temp.name, 'rb'), as_attachment=True, filename='actividades.xlsx')

    # Delete the temporary file
    os.unlink(temp.name)

    return response


# abre el archivo de actividades en google sheet
@login_required
# @entidad_requerida
def cronograma(request):
    context = {
        'cronograma_url': 'https://docs.google.com/spreadsheets/d/1rCKLRV191H3TI5yXis7XO3GS9OCXDWohL24ZgfNQGS8/edit',
    }

    return render(request, 'cronograma.html', context)



# abre el archivo de resumen de tiempo, horas y estado de papeles de trabajo
@login_required
def resumen_tiempo(request):
    context ={
        'resumen_tiempo_url': 'https://docs.google.com/spreadsheets/d/17DIx4ImW3199ZLz0AUV0FDWrYfnkUnIJSuwWDxbOkZs/edit',
        'resumen_horas_url': 'https://docs.google.com/spreadsheets/d/1Zf_uT8ugjSuX7q8TY1lFo2dk7kLQKSsWx38Vl71l0So/edit',
        'estado_papeles_url': 'https://docs.google.com/spreadsheets/d/1fnNZ5ltcC1xELHq2n2-ZvDaJL2k1kd2i5Vun46EODc0/edit',
    }
    return render(request, 'resumen_tiempo.html', context)

@login_required
@entidad_requerida
def marcas(request):
    context = urls["marcas"]
    return render(request, 'marcas.html', context)

@login_required
@entidad_requerida
def monedas(request):
    context = urls["monedas"]
    return render(request, 'monedas.html', context)